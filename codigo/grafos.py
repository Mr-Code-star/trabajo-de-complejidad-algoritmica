import os
import matplotlib.pyplot as plt
from collections import defaultdict, deque
from openpyxl import load_workbook, Workbook
import math
import tkinter as tk
from tkinter import Canvas
import random


# Configuracion de rutas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATASET_DIR = os.path.join(BASE_DIR, "../Dataset")

# UFDS para comunidades
class UFDS:
    def __init__(self):
        self.parent = {}
        self.rank = {}
    
    def make_set(self, x):
        if x not in self.parent:
            self.parent[x] = x
            self.rank[x] = 0
    
    def find(self, x):
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]
    
    def union(self, x, y):
        rx = self.find(x)
        ry = self.find(y)
        
        if rx == ry:
            return
        
        if self.rank[rx] < self.rank[ry]:
            self.parent[rx] = ry
        elif self.rank[rx] > self.rank[ry]:
            self.parent[ry] = rx
        else:
            self.parent[ry] = rx
            self.rank[rx] += 1


# Utilidades internas
def _abrir_hoja(path_xlsx: str):
    #ABRE ARCHIVO EXCEL
    archivo = os.path.join(DATASET_DIR, path_xlsx)
    if not os.path.exists(archivo):
        raise FileNotFoundError(
            f"No se encontro '{path_xlsx}' en la carpeta Dataset.\n"
            f"Ruta esperada: {archivo}"
        )
    wb = load_workbook(archivo, data_only=True)
    return wb.active


# Cargar usuarios y posts desde usuarios.xlsx
def cargar_usuarios():
    usuarios = {}
    posts = {}

    ws = _abrir_hoja("usuarios.xlsx")

    for id_, nombre, post in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if id_ is None:
            continue
        sid = str(id_).strip()
        usuarios[sid] = (nombre or "").strip()
        posts[sid] = (post or "").strip()

    return usuarios, posts

def cargar_posts():
    """
    Carga todos los posts desde posts.xlsx.
    Retorna:
      - posts_por_id: {id_post(int): {"id_usuario": str, "contenido": str}}
      - posts_por_usuario: {id_usuario(str): [id_post1, id_post2, ...]}
    """
    posts_por_id = {}
    posts_por_usuario = defaultdict(list)

    try:
        ws = _abrir_hoja("posts.xlsx")
    except FileNotFoundError:
        # Si no existe el archivo, no hay posts aún
        return posts_por_id, posts_por_usuario

    for id_post, id_usuario, contenido in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if id_post is None or id_usuario is None:
            continue

        pid = int(id_post)
        uid = str(id_usuario).strip()
        texto = (contenido or "").strip()

        posts_por_id[pid] = {
            "id_usuario": uid,
            "contenido": texto
        }
        posts_por_usuario[uid].append(pid)

    return posts_por_id, posts_por_usuario


def crear_post(id_usuario, contenido):
    """
    Crea un nuevo post para el usuario dado y lo guarda en posts.xlsx.
    Retorna el nuevo id_post (int).
    """
    archivo = os.path.join(DATASET_DIR, "posts.xlsx")
    id_usuario = str(id_usuario).strip()

    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active

        max_id = 0
        for id_post, _, _ in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if id_post is None:
                continue
            try:
                max_id = max(max_id, int(id_post))
            except ValueError:
                continue

        nuevo_id = max_id + 1
    else:
        # Crear archivo desde cero
        wb = Workbook()
        ws = wb.active
        ws.title = "Posts"
        ws.append(["id_post", "id_usuario", "contenido"])
        nuevo_id = 1

    ws.append([nuevo_id, id_usuario, contenido])
    wb.save(archivo)
    return nuevo_id


def actualizar_post_usuario(id_usuario, nuevo_post):
    """
    Actualiza el contenido del post de un usuario en usuarios.xlsx.
    Retorna (exito: bool, mensaje: str)
    """
    archivo = os.path.join(DATASET_DIR, "usuarios.xlsx")
    
    if not os.path.exists(archivo):
        return False, "No se encontró el archivo usuarios.xlsx"
    
    wb = load_workbook(archivo)
    ws = wb.active
    
    id_str = str(id_usuario).strip()
    encontrado = False
    
    # Se asume formato: [id, nombre, post] desde la fila 2
    for row in ws.iter_rows(min_row=2, max_col=3):
        cell_id = row[0].value
        if cell_id is None:
            continue
        
        if str(cell_id).strip() == id_str:
            # Columna 3 = post
            row[2].value = nuevo_post
            encontrado = True
            break
    
    if not encontrado:
        return False, f"No se encontró el usuario con id {id_str}"
    
    wb.save(archivo)
    return True, "Post actualizado correctamente."


# Cargar amistades desde amistades.xlsx
def cargar_grafo():
    grafo = defaultdict(list)

    ws = _abrir_hoja("amistades.xlsx")

    for id1, id2 in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if id1 is None or id2 is None:
            continue
        a, b = str(id1).strip(), str(id2).strip()
        if a == "" or b == "":
            continue
        grafo[a].append(b)
        grafo[b].append(a)  # no dirigido

    return grafo


# Cargar comunidades desde comunidades.xlsx
def cargar_comunidades():
    #Cargar comunidades desde Excel.
    comunidades = defaultdict(list)  # id_comunidad -> [usuarios]
    nombres_comunidades = {}  # id_comunidad -> nombre
    usuario_comunidad = {}  # id_usuario -> id_comunidad
    
    try:
        ws = _abrir_hoja("comunidades.xlsx")
        
        for id_com, nombre_com, id_user in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if id_com is None or id_user is None:
                continue
            
            id_com_str = str(id_com).strip()
            id_user_str = str(id_user).strip()
            
            if nombre_com:
                nombres_comunidades[id_com_str] = str(nombre_com).strip()
            
            comunidades[id_com_str].append(id_user_str)
            usuario_comunidad[id_user_str] = id_com_str
            
    except FileNotFoundError:
        # Si el archivo no existe se crea cuando se agregue una comunidad
        pass
    
    return comunidades, nombres_comunidades, usuario_comunidad

# Cargar likes desde likes.xlsx
def cargar_likes():
    """
    Carga los likes desde el archivo likes.xlsx.
    Retorna una lista de diccionarios:
    [{'id_like': int, 'id_usuario_like': 'id', 'id_post': int}, ...]
    """
    likes = []

    try:
        ws = _abrir_hoja("likes.xlsx")
    except FileNotFoundError:
        # Si no existe el archivo, no hay likes aún
        return likes

    for id_like, id_usuario_like, id_post in ws.iter_rows(
            min_row=2, max_col=3, values_only=True):

        if id_like is None or id_usuario_like is None or id_post is None:
            continue

        likes.append({
            "id_like": int(id_like),
            "id_usuario_like": str(id_usuario_like).strip(),
            "id_post": int(id_post)
        })

    return likes


# Registrar like
def registrar_like(id_usuario_like, id_post):
    """
    Registra un nuevo like en likes.xlsx.
    - Evita likes duplicados del mismo usuario al mismo post.
    - Crea el archivo si no existe.

    Retorna: (exito: bool, mensaje: str)
    """
    archivo = os.path.join(DATASET_DIR, "likes.xlsx")

    id_usuario_like = str(id_usuario_like).strip()
    id_post = int(id_post)

    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active

        pares_existentes = set()
        max_id = 0

        for id_like, u_like, p_id in ws.iter_rows(
                min_row=2, max_col=3, values_only=True):

            if id_like is None or u_like is None or p_id is None:
                continue

            u_like_str = str(u_like).strip()
            try:
                p_id_int = int(p_id)
            except ValueError:
                continue

            pares_existentes.add((u_like_str, p_id_int))

            try:
                max_id = max(max_id, int(id_like))
            except ValueError:
                continue

        if (id_usuario_like, id_post) in pares_existentes:
            return False, "El usuario ya dio like a este post."

        nuevo_id = max_id + 1

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Likes"
        ws.append(["id_like", "id_usuario_like", "id_post"])
        nuevo_id = 1

    ws.append([nuevo_id, id_usuario_like, id_post])
    wb.save(archivo)

    return True, "Like registrado correctamente."


# Contar likes por post
def contar_likes_por_post(likes):
    """
    Recibe la lista de likes (salida de cargar_likes)
    y devuelve un diccionario: {id_post: cantidad_likes}
    """
    conteo = defaultdict(int)

    for row in likes:
        post_id = row["id_post"]
        conteo[post_id] += 1

    return conteo


# Divide y Venceras
def max_post_por_likes_divide_venceras(items):
    """
    items: lista de tuplas (id_post, likes)
    Retorna la tupla (id_post, likes) con mayor cantidad de likes,
    usando la estrategia Divide y Vencerás.
    """
    if not items:
        return None
    
    # Caso base: un solo elemento
    if len(items) == 1:
        return items[0]
    
    # Dividir en dos mitades
    mid = len(items) // 2
    izquierda = items[:mid]
    derecha = items[mid:]
    
    # Resolver recursivamente
    max_izq = max_post_por_likes_divide_venceras(izquierda)
    max_der = max_post_por_likes_divide_venceras(derecha)
    
    # Combinar: quedarse con el que tiene más likes
    if max_izq[1] >= max_der[1]:
        return max_izq
    else:
        return max_der

def merge_sort_posts_por_likes(items):
    """
    Ordena una lista de tuplas (id_post, likes) de MAYOR a MENOR likes
    usando MergeSort (Divide y Vencerás).
    """
    if len(items) <= 1:
        return items
    
    mid = len(items) // 2
    izquierda = merge_sort_posts_por_likes(items[:mid])
    derecha = merge_sort_posts_por_likes(items[mid:])
    
    return _merge_por_likes(izquierda, derecha)

def _merge_por_likes(izquierda, derecha):
    """
    Fase de combinación de MergeSort.
    Combina dos listas ya ordenadas por likes (descendente).
    """
    resultado = []
    i = j = 0
    
    while i < len(izquierda) and j < len(derecha):
        # >= para que quede de mayor a menor
        if izquierda[i][1] >= derecha[j][1]:
            resultado.append(izquierda[i])
            i += 1
        else:
            resultado.append(derecha[j])
            j += 1
    
    # Agregar lo que falte
    while i < len(izquierda):
        resultado.append(izquierda[i])
        i += 1
    
    while j < len(derecha):
        resultado.append(derecha[j])
        j += 1
    
    return resultado

def obtener_top_posts(likes, k=5):
    """
    Retorna una lista con los k posts más populares:
    [(id_post, likes), ...] ya ordenados de mayor a menor.
    """
    conteo = contar_likes_por_post(likes)
    items = list(conteo.items())
    
    if not items:
        return []
    
    ordenados = merge_sort_posts_por_likes(items)
    return ordenados[:k]


# Guardar comunidades en Excel
def guardar_comunidades(comunidades, nombres_comunidades):
    #Guarda comunidades en Excel
    archivo = os.path.join(DATASET_DIR, "comunidades.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Comunidades"
    
    # Encabezados
    ws.append(["id_comunidad", "nombre_comunidad", "id_usuario"])
    
    # Datos
    for id_com, usuarios in comunidades.items():
        nombre_com = nombres_comunidades.get(id_com, f"Comunidad {id_com}")
        for usuario in usuarios:
            ws.append([id_com, nombre_com, usuario])
    
    wb.save(archivo)


# Sistema de comunidades con UFDS
class SistemaComunidades:
    def __init__(self):
        self.ufds = UFDS()
        self.comunidades, self.nombres_comunidades, self.usuario_comunidad = cargar_comunidades()
        self._inicializar_ufds()
    
    def _inicializar_ufds(self):
        #Inicializa UFDS con los usuarios de las comunidades
        for id_com, usuarios in self.comunidades.items():
            for usuario in usuarios:
                self.ufds.make_set(usuario)
                if len(usuarios) > 1:
                    # Unir todos los usuarios de la misma comunidad
                    primer_usuario = usuarios[0]
                    self.ufds.union(primer_usuario, usuario)
    
    def crear_comunidad(self, nombre_comunidad, usuarios):
        #Crear una nueva comunidad
        if not usuarios:
            return False, "Debe incluir al menos un usuario"
        
        # Generar nuevo ID de comunidad
        nuevo_id = str(max([int(k) for k in self.comunidades.keys()] + [0]) + 1)
        
        # Agregar usuarios a la comunidad
        self.comunidades[nuevo_id] = usuarios
        self.nombres_comunidades[nuevo_id] = nombre_comunidad
        
        # Actualizar UFDS y mapeo de usuario a comunidad
        for usuario in usuarios:
            self.ufds.make_set(usuario)
            self.usuario_comunidad[usuario] = nuevo_id
            
            # Unir usuarios de la misma comunidad
            if usuario != usuarios[0]:
                self.ufds.union(usuarios[0], usuario)
        
        # Guardar en Excel
        guardar_comunidades(self.comunidades, self.nombres_comunidades)
        
        return True, f"Comunidad '{nombre_comunidad}' creada exitosamente"
    
    def obtener_comunidad_usuario(self, usuario_id):
        #Obtiene la comunidad de un usuario
        return self.usuario_comunidad.get(usuario_id)
    
    def obtener_usuarios_comunidad(self, comunidad_id):
        #Obtiene todos los usuarios de una comunidad
        return self.comunidades.get(comunidad_id, [])
    
    def obtener_todas_comunidades(self, usuarios_dict):
        #Obtiene todas las comunidades con sus nombres y usuarios
        resultado = []
        for id_com, usuarios_lista in self.comunidades.items():
            nombre = self.nombres_comunidades.get(id_com, f"Comunidad {id_com}")
            nombres_usuarios = [usuarios_dict.get(user_id, f"Usuario {user_id}") 
                              for user_id in usuarios_lista]
            resultado.append({
                'id': id_com,
                'nombre': nombre,
                'usuarios': usuarios_lista,
                'nombres_usuarios': nombres_usuarios
            })
        return resultado
    
    def son_misma_comunidad(self, usuario1, usuario2):
        #Verifica si dos usuarios estan en la misma comunidad
        if usuario1 not in self.ufds.parent or usuario2 not in self.ufds.parent:
            return False
        return self.ufds.find(usuario1) == self.ufds.find(usuario2)


# BFS para el camino mas corto
def camino_mas_corto(grafo, inicio, destino):
    #Encuentra el camino mas corto usando BFS
    cola = deque([(inicio, [inicio])])
    visitados = set([inicio])

    while cola:
        nodo, camino = cola.popleft()
        if nodo == destino:
            return camino
        for vecino in grafo[nodo]:
            if vecino not in visitados:
                visitados.add(vecino)
                cola.append((vecino, camino + [vecino]))
    return None


# Recomendacion de amigos
def recomendar_amigos(grafo, usuario):
    #Devuelve amigos de amigos que no sean ya amigos directos
    directos = set(grafo[usuario])
    sugerencias = set()
    for amigo in directos:
        for amigo_de_amigo in grafo[amigo]:
            if amigo_de_amigo != usuario and amigo_de_amigo not in directos:
                sugerencias.add(amigo_de_amigo)
    return list(sugerencias)


# Calcular grado de cada nodo
def calcular_grados(grafo):
    #Calcula el grado de cada nodo en el grafo
    grados = {}
    for nodo, vecinos in grafo.items():
        grados[nodo] = len(vecinos)
    # Asegurar que todos los nodos aislados tengan grado 0
    for nodo in grafo:
        if nodo not in grados:
            grados[nodo] = 0
    return grados


# Obtener subgrafo relevante (vecindario de N saltos)
def obtener_subgrafo(grafo, nodos_centrales, saltos=2):
    
    #Obtiene un subgrafo que incluye los nodos centrales y sus vecinos hasta N saltos.
    
    #Args:
    #    grafo: grafo completo
    #    nodos_centrales: lista de nodos centrales para el subgrafo
    #    saltos: numero de saltos desde los nodos centrales
    
    #Returns:
    #    subgrafo: diccionario con solo los nodos relevantes
    #    nodos_incluidos: conjunto de nodos en el subgrafo
    
    nodos_incluidos = set(nodos_centrales)
    nodos_por_explorar = set(nodos_centrales)
    
    for _ in range(saltos):
        nuevos_nodos = set()
        for nodo in nodos_por_explorar:
            if nodo in grafo:
                vecinos = grafo[nodo]
                nuevos_nodos.update(vecinos)
        nodos_incluidos.update(nuevos_nodos)
        nodos_por_explorar = nuevos_nodos - nodos_incluidos
    
    # Crear subgrafo
    subgrafo = defaultdict(list)
    for nodo in nodos_incluidos:
        if nodo in grafo:
            for vecino in grafo[nodo]:
                if vecino in nodos_incluidos:
                    subgrafo[nodo].append(vecino)
    
    return subgrafo, nodos_incluidos


# Clase para visualizacion interactiva del grafo
class VisualizadorGrafo:
    def __init__(self, canvas, grafo, usuarios, ancho=800, alto=600):
        self.canvas = canvas
        self.grafo = grafo
        self.usuarios = usuarios
        self.ancho = ancho
        self.alto = alto
        self.posiciones = {}
        self.nodos_dibujados = {}
        self.aristas_dibujadas = []
        self.zoom = 1.0
        self.offset_x = 0
        self.offset_y = 0
        
    def limpiar(self):
        #Limpia el canvas
        self.canvas.delete("all")
        self.nodos_dibujados = {}
        self.aristas_dibujadas = []
        
    def calcular_layout_fuerza(self, nodos, iteraciones=50):
        """
        Calcula las posiciones usando un algoritmo de fuerza simplificado.
        """
        # Limitar número de nodos para mejor rendimiento
        num_nodos = len(nodos)
    
        # Ajustar iteraciones según el tamaño
        if num_nodos > 20:
            iteraciones = min(30, iteraciones)
        elif num_nodos > 10:
            iteraciones = min(40, iteraciones)
    
        # Inicializar posiciones aleatoriamente
        posiciones = {}
        for nodo in nodos:
            posiciones[nodo] = [
            random.uniform(100, self.ancho - 100),
            random.uniform(100, self.alto - 100)
        ]
    
        # Si hay pocos nodos, usar layout circular simple
        if num_nodos <= 5:  # Cambiado de 3 a 5
            return self._layout_circular(list(nodos))
        
        # Parametros del algoritmo
        k = math.sqrt((self.ancho * self.alto) / num_nodos) if num_nodos > 0 else 100
        c_rep = k * k  # Constante de repulsion
        c_spring = 1.0  # Constante del resorte
        damping = 0.9  # Factor de amortiguacion
        
        nodos_lista = list(nodos)
        
        for iteracion in range(iteraciones):
            # Reducir fuerzas en cada iteración para convergencia más rápida
            factor = 1.0 - (iteracion / iteraciones) * 0.5
            
            fuerzas = {nodo: [0, 0] for nodo in nodos}
            
            # Fuerzas de repulsion (optimizado)
            for i, nodo1 in enumerate(nodos_lista):
                for nodo2 in nodos_lista[i+1:]:
                    dx = posiciones[nodo2][0] - posiciones[nodo1][0]
                    dy = posiciones[nodo2][1] - posiciones[nodo1][1]
                    dist = math.sqrt(dx*dx + dy*dy)
                    
                    if dist > 0 and dist < k * 3:  # Solo calcular si están cerca
                        # Fuerza de repulsion
                        f_rep = (c_rep / (dist * dist)) * factor
                        fx = (dx / dist) * f_rep
                        fy = (dy / dist) * f_rep
                        
                        fuerzas[nodo1][0] -= fx
                        fuerzas[nodo1][1] -= fy
                        fuerzas[nodo2][0] += fx
                        fuerzas[nodo2][1] += fy
            
            # Fuerzas de atraccion para nodos conectados
            for nodo in nodos:
                if nodo in self.grafo:
                    for vecino in self.grafo[nodo]:
                        if vecino in nodos:
                            dx = posiciones[vecino][0] - posiciones[nodo][0]
                            dy = posiciones[vecino][1] - posiciones[nodo][1]
                            dist = math.sqrt(dx*dx + dy*dy)
                            
                            if dist > 0:
                                # Fuerza del resorte
                                f_spring = c_spring * math.log(dist / k) * factor
                                fx = (dx / dist) * f_spring
                                fy = (dy / dist) * f_spring
                                
                                fuerzas[nodo][0] += fx * 0.5
                                fuerzas[nodo][1] += fy * 0.5
            
            # Actualizar posiciones
            for nodo in nodos:
                posiciones[nodo][0] += fuerzas[nodo][0] * damping
                posiciones[nodo][1] += fuerzas[nodo][1] * damping
                
                # Mantener dentro de los limites
                posiciones[nodo][0] = max(50, min(self.ancho - 50, posiciones[nodo][0]))
                posiciones[nodo][1] = max(50, min(self.alto - 50, posiciones[nodo][1]))
        
        return posiciones
    
    
    def _layout_circular(self, nodos):
        """Layout circular para pocos nodos"""
        posiciones = {}
        n = len(nodos)
    
        if n == 0:
            return posiciones
    
        # Centro del canvas
        cx = self.ancho / 2
        cy = self.alto / 2
    
        # Radio del círculo - más grande para mejor visualización
        radio = min(self.ancho, self.alto) * 0.25
    
        if n == 1:
            posiciones[nodos[0]] = [cx, cy]
        elif n == 2:
            # Para 2 nodos, ponerlos horizontalmente
            posiciones[nodos[0]] = [cx - radio, cy]
            posiciones[nodos[1]] = [cx + radio, cy]
        else:
            for i, nodo in enumerate(nodos):
                angulo = 2 * math.pi * i / n - math.pi / 2  # Empezar desde arriba
                x = cx + radio * math.cos(angulo)
                y = cy + radio * math.sin(angulo)
                posiciones[nodo] = [x, y]
        return posiciones
    
    def dibujar_grafo(self, subgrafo=None, camino=None, nodos_destacados=None):
        """
        Dibuja el grafo en el canvas.
        
        Args:
            subgrafo: si se proporciona, solo dibuja este subgrafo
            camino: lista de nodos que forman un camino para destacar
            nodos_destacados: conjunto de nodos para destacar
        """
        self.limpiar()
        
        grafo_a_dibujar = subgrafo if subgrafo else self.grafo
        
        # Obtener todos los nodos del grafo
        nodos = set()
        for nodo in grafo_a_dibujar:
            nodos.add(nodo)
            nodos.update(grafo_a_dibujar[nodo])
        
        # IMPORTANTE: Si se especifica nodos_destacados, SOLO usar esos nodos
        if nodos_destacados:
            nodos = set(nodos_destacados)
        
        if not nodos:
            return
        
        # Calcular posiciones
        self.posiciones = self.calcular_layout_fuerza(nodos)
        
        # Dibujar aristas primero para que queden debajo de los nodos
        aristas_dibujadas = set()
        for nodo in grafo_a_dibujar:
            if nodo not in nodos:  # Solo dibujar si el nodo está en el conjunto
                continue
            for vecino in grafo_a_dibujar[nodo]:
                if vecino not in nodos:  # Solo dibujar si el vecino está en el conjunto
                    continue
                arista = tuple(sorted([nodo, vecino]))
                if arista not in aristas_dibujadas and vecino in self.posiciones:
                    aristas_dibujadas.add(arista)
                    
                    x1, y1 = self.posiciones[nodo]
                    x2, y2 = self.posiciones[vecino]
                    
                    # Determinar color de la arista
                    color = "#cccccc"
                    ancho = 1
                    
                    if camino:
                        # Verificar si la arista esta en el camino
                        for i in range(len(camino) - 1):
                            if (camino[i] == nodo and camino[i+1] == vecino) or \
                               (camino[i] == vecino and camino[i+1] == nodo):
                                color = "#ff4444"
                                ancho = 3
                                break
                    
                    linea = self.canvas.create_line(
                        x1, y1, x2, y2,
                        fill=color, width=ancho, tags="arista"
                    )
                    self.aristas_dibujadas.append(linea)
        
        # Dibujar nodos
        for nodo in nodos:
            if nodo in self.posiciones:
                x, y = self.posiciones[nodo]
                
                # Determinar color y tamaño del nodo
                color = "#64b5f6"
                radio = 25  # Aumentado de 20 a 25
                color_texto = "black"
                
                if camino:
                    if nodo == camino[0]:  # Origen
                        color = "#4caf50"
                        radio = 30  # Aumentado de 25 a 30
                        color_texto = "white"
                    elif nodo == camino[-1]:  # Destino
                        color = "#f44336"
                        radio = 30
                        color_texto = "white"
                    elif nodo in camino:  # En el camino
                        color = "#ffd54f"
                        radio = 27
                
                if nodos_destacados and nodo in nodos_destacados:
                    radio = 30  # Aumentado de 25 a 30
                    if not camino or nodo not in camino:
                        color = "#9c27b0"
                        color_texto = "white"
                
                # Dibujar circulo del nodo
                circulo = self.canvas.create_oval(
                    x - radio, y - radio,
                    x + radio, y + radio,
                    fill=color, outline="white", width=3,  # Borde más grueso
                    tags=("nodo", f"nodo_{nodo}")
                )
                
                # Dibujar nombre del nodo
                nombre = self.usuarios.get(nodo, f"ID: {nodo}")
                # Truncar nombre si es muy largo
                if len(nombre) > 10:
                    nombre = nombre[:8] + "..."
                
                texto = self.canvas.create_text(
                    x, y,
                    text=nombre,
                    font=("Arial", 10, "bold"),  # Aumentado de 9 a 10
                    fill=color_texto,
                    tags=("texto", f"texto_{nodo}")
                )
                
                self.nodos_dibujados[nodo] = (circulo, texto)
                
                # Agregar tooltip con nombre completo
                self.agregar_tooltip(nodo, circulo)
                
    
    def agregar_tooltip(self, nodo_id, elemento):
        #Agrega un tooltip al pasar el mouse sobre un nodo
        nombre_completo = self.usuarios.get(nodo_id, f"Usuario {nodo_id}")
        grado = len(self.grafo.get(nodo_id, []))
        
        def mostrar_tooltip(event):
            x = self.canvas.winfo_pointerx() - self.canvas.winfo_rootx() + 10
            y = self.canvas.winfo_pointery() - self.canvas.winfo_rooty() - 30
            
            self.canvas.delete("tooltip")
            
            # Crear tooltip
            rect = self.canvas.create_rectangle(
                x, y, x + 150, y + 40,
                fill="#333333", outline="#555555",
                tags="tooltip"
            )
            
            texto = f"{nombre_completo}\nConexiones: {grado}"
            self.canvas.create_text(
                x + 75, y + 20,
                text=texto,
                fill="white",
                font=("Arial", 9),
                tags="tooltip"
            )
        
        def ocultar_tooltip(event):
            self.canvas.delete("tooltip")
        
        self.canvas.tag_bind(elemento, "<Enter>", mostrar_tooltip)
        self.canvas.tag_bind(elemento, "<Leave>", ocultar_tooltip)


# Analisis del grafo
def analizar_grafo(grafo, usuarios):
   
    #Analiza las propiedades del grafo y devuelve estadisticas 
    
    # Calcular grados
    grados = calcular_grados(grafo)
    
    # Estadisticas basicas
    num_nodos = len(usuarios)
    num_aristas = sum(len(vecinos) for vecinos in grafo.values()) // 2
    
    if grados:
        grado_promedio = sum(grados.values()) / len(grados)
        grado_max = max(grados.values())
        grado_min = min(grados.values())
        
        # Encontrar nodos
        nodos_mas_conectados = sorted(grados.items(), 
                                     key=lambda x: x[1], 
                                     reverse=True)[:5]
    else:
        grado_promedio = 0
        grado_max = 0
        grado_min = 0
        nodos_mas_conectados = []
    
    # Calcular densidad
    densidad = (2 * num_aristas) / (num_nodos * (num_nodos - 1)) if num_nodos > 1 else 0
    
    return {
        'num_nodos': num_nodos,
        'num_aristas': num_aristas,
        'grado_promedio': grado_promedio,
        'grado_max': grado_max,
        'grado_min': grado_min,
        'densidad': densidad,
        'nodos_mas_conectados': nodos_mas_conectados
    }

def obtener_subgrafo_comunidad(grafo, id_comunidad, comunidades):
    """
    Obtiene el subgrafo de una comunidad específica.
    
    Args:
        grafo: grafo completo
        id_comunidad: ID de la comunidad a visualizar
        comunidades: diccionario de comunidades {id_com: [usuarios]}
    
    Returns:
        subgrafo: diccionario con solo los nodos y aristas de la comunidad
        nodos_incluidos: conjunto de nodos en la comunidad
    """
    if id_comunidad not in comunidades:
        return defaultdict(list), set()
    
    nodos_comunidad = set(comunidades[id_comunidad])
    
    # Crear subgrafo solo con conexiones dentro de la comunidad
    subgrafo = defaultdict(list)
    for nodo in nodos_comunidad:
        if nodo in grafo:
            for vecino in grafo[nodo]:
                if vecino in nodos_comunidad:
                    # Evitar duplicados
                    if vecino not in subgrafo[nodo]:
                        subgrafo[nodo].append(vecino)
    
    return subgrafo, nodos_comunidad

# Funciones auxiliares para visualizacion simplificada (mantener compatibilidad)
def dibujar_grafo_graphviz(*args, **kwargs):
    #Funcion mantenida por compatibilidad pero ya no usada
    return "Visualizacion integrada en la interfaz"

def dibujar_grafo_comunidades(*args, **kwargs):
    #Funcion mantenida por compatibilidad pero ya no usada
    return "Visualizacion integrada en la interfaz"
